"""Extract a small, local portion of the ASHRAE HOF 2025 design-conditions
spreadsheet for use as a test fixture.

The full spreadsheet (``HOF_2025_Climate_Design_Conditions_SI.xlsx``) is ~45 MB
and lives outside this repository (on a shared drive). The tests in this folder
must run offline and be self-contained, so this script pulls just the two rows we
care about (Chicago O'Hare and Glasgow Bishopton) plus the column header map and
writes them to ``tools/testdata/ashrae_hof2025_design_extract.json``.

Run it once (when the source spreadsheet is available) to regenerate the fixture::

    python tools/extract_ashrae_reference.py

Values are kept verbatim in the spreadsheet's "SI" units (degrees C, kJ/kg, m/s,
degrees, %, mm) -- *not* base SI. The DDY / EPW design conditions are published in
exactly these units, so the test can compare them directly without conversion.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# Default location of the source spreadsheet (shared drive, not in the repo).
SOURCE_XLSX = Path(
    "/Users/prastogi/Library/CloudStorage/OneDrive-SharedLibraries-GRESBB.V/"
    "Shared Files - 22 Innovation/Data/climate/ashrae/"
    "HOF_2025_Climate_Design_Conditions_SI.xlsx"
)

OUT_PATH = Path(__file__).resolve().parent / "testdata" / "ashrae_hof2025_design_extract.json"

# 1-based worksheet rows for the two stations we mirror in the examples.
STATION_ROWS = {
    "chicago": 9255,  # USA IL CHICAGO O'HARE, WMO 725300
    "glasgow": 839,   # GBR GLASGOW BISHOPTON, WMO 031340
}

# We keep columns A..CL (1..90): station information (A..O) plus all the annual
# heating / cooling / extreme design conditions that appear in the EPW header line
# and the DDY design-day objects. Columns beyond CL (monthly tables, degree days,
# precipitation, solar, trends) are not carried in the EPW/DDY design conditions.
LAST_COL = "CL"


def _ffill(row):
    out, last = [], None
    for v in row:
        if v is not None and str(v).strip() != "":
            last = v
        out.append(last)
    return out


def build_extract(xlsx_path: Path = SOURCE_XLSX) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Stations"]

    header_rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=4, values_only=True)]
    section = _ffill(header_rows[0])
    group = _ffill(header_rows[1])
    sub1 = header_rows[2]
    sub2 = header_rows[3]

    last_idx = column_index_from_string(LAST_COL)  # 90

    columns = {}
    for ci in range(1, last_idx + 1):
        letter = get_column_letter(ci)
        parts = [
            str(x).strip()
            for x in (section[ci - 1], group[ci - 1], sub1[ci - 1], sub2[ci - 1])
            if x is not None and str(x).strip() != ""
        ]
        columns[letter] = " | ".join(parts)

    stations = {}
    for name, rownum in STATION_ROWS.items():
        row = list(next(ws.iter_rows(min_row=rownum, max_row=rownum, values_only=True)))
        values = {}
        for ci in range(1, last_idx + 1):
            letter = get_column_letter(ci)
            v = row[ci - 1]
            # Normalise integer-valued floats (e.g. wind direction 280.0 -> 280).
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            values[letter] = v
        stations[name] = {"row": rownum, "values": values}

    return {
        "_source": f"{xlsx_path.name}, sheet 'Stations', header rows 1-4",
        "_units": "ASHRAE SI sheet units: temperatures in degC, enthalpy in kJ/kg, "
        "wind speed in m/s, wind direction in degrees, humidity ratio in g/kg, "
        "precipitation in mm. NOT base SI.",
        "_columns_covered": f"A..{LAST_COL}",
        "columns": columns,
        "stations": stations,
    }


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(
            f"Source spreadsheet not found:\n  {SOURCE_XLSX}\n"
            "This script only needs to run when regenerating the fixture."
        )
    extract = build_extract()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(extract, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
